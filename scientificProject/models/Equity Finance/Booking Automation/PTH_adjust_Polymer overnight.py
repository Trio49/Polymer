# -*- coding: utf-8 -*-
"""
Created on Mon Mar 22 2020

@author: tyin
"""


# Last edited by:
# Zach Zhang 06222021
# PTH update model  03222020 
import pytz
import pandas as pd
import numpy as np
#import matplotlib.pyplot as plt 
from datetime import datetime, timedelta
from os import listdir

pd.set_option('display.max_columns', 20)


today=datetime.now(pytz.timezone('Asia/Shanghai'))
# =============================================================================
# # =============================================================================
today1=today.date()-timedelta(days=1)

# today=today.date()-timedelta(days=1)


todaydate=today.strftime('%Y%m%d')
today1=today1.strftime('%Y%m%d')

def firm_positions():
# read the FIRM FLAT positions from some place!
    firm_positions_file_path="//paghk.local/shares/Enfusion/Polymer/"+today1
    firm_positions_list_dir=listdir(firm_positions_file_path)
    
    firm_positions_list_dir = firm_positions_list_dir[::-1]
    for file_name in firm_positions_list_dir:
        if "Firm_Flat_All_Pos" in file_name:
            firm_positions_file_path=firm_positions_file_path+"\\"+file_name
            break
    
    ####tyin
    firm_positions = pd.read_csv(firm_positions_file_path)
    #firm_positions = pd.read_csv("\\paghk.local\shares\Enfusion\Polymer\" + today)
#    
 #firm_positions_path=r"C:\Users\tyin\Documents\report_FIRM_FLAT_TYIN0322.csv"
#    firm_positions=pd.read_csv(firm_positions_path)
    firm_positions=firm_positions.iloc[1:]
    firm_positions.describe()
    firm_positions=firm_positions[firm_positions["Book Name"].str.contains("PTH")]
    firm_positions=firm_positions[firm_positions["Position"]!=0]
    firm_positions=firm_positions[["Sub Business Unit",
                                  "Book Name",
                                  "DW Underlying Ticker",
                                  "Description",
                                  "Position",
                                  "Custodian Account Display Name",
                                  "Last Trade Date"]]
    
    #to the sub-PM level to match with trade blotter strategy column 
    firm_positions["strats"]=firm_positions["Book Name"].apply(lambda x:x.split("(")[1].replace(")",""))
    
    #firm_positions
    
    
    firm_positions["Custodian_match"]=firm_positions["Custodian Account Display Name"].apply(lambda x:tuple(x.split("_")[1:3]))
    return(firm_positions)
                
    


firm_positions= firm_positions()
firm_positions["ticker_PM"]= firm_positions["DW Underlying Ticker"]+" "+ firm_positions["strats"]

ticker_PM_list=firm_positions["ticker_PM"].tolist()
######################################################################## list all existing PM positions for those PTH names



firm_positions_file_path1=r"\\paghk.local\shares\Enfusion\Polymer\\"+today
firm_positions_list_dir1=listdir(firm_positions_file_path1)
for file_name in firm_positions_list_dir1:
    if "Firm_Flat_All_Pos" in file_name:
        firm_positions_file_path1=firm_positions_file_path1+"\\"+file_name
        break
    

firm_positions_new = pd.read_csv(firm_positions_file_path1)
#    firm_positions_path=r"C:\Users\tyin\Documents\report_FIRM_FLAT_TYIN0322.csv"
#    firm_positions=pd.read_csv(firm_positions_path)
firm_positions_new=firm_positions_new.iloc[1:]
firm_positions_new=firm_positions_new[firm_positions_new["Position"]!=0]


firm_positions_new=firm_positions_new[~firm_positions_new["Book Name"].str.contains("PTH")]


firm_positions_new=firm_positions_new[["Sub Business Unit",
                                  "Book Name",
                                  "DW Underlying Ticker",
                                  "Description",
                                  "Position",
                                  "Custodian Account Display Name",
                                  "Last Trade Date"]]
    
    #to the sub-PM level to match with trade blotter strategy column 
firm_positions_new["strats"]=firm_positions_new["Book Name"].apply(lambda x:x[0:5])
    
    #firm_positions
    
    
firm_positions_new["Custodian_match"]=firm_positions_new["Custodian Account Display Name"].apply(lambda x:tuple(x.split("_")[1:3]))

firm_positions_new["ticker_PM"]=firm_positions_new["DW Underlying Ticker"]+" "+firm_positions_new["strats"]


firm_positions_new_=firm_positions_new[firm_positions_new["ticker_PM"].isin(ticker_PM_list)]


whole = firm_positions.append(firm_positions_new_)





########################################################################




# read the REPORT_Polymer_Blotter_Tony_202
def Trade_Blotter():
     #Trade_Blotter_file_path=r"\\paghk.local\shares\Enfusion\Polymer\\"+today
    Trade_Blotter_file_path=r"\\paghk.local\shares\Enfusion\Polymer\\"+today
    list_dir=listdir(Trade_Blotter_file_path)
    
  
    
    list_dir = list_dir[::-1]
    
    for file_name in list_dir:
        if "Polymer_Trade Blotter - PM ALL Overnight (New)" in file_name:
            Trade_Blotter_file_path=Trade_Blotter_file_path+"\\"+file_name
            break
    
    Trade_Blotter = pd.read_csv(Trade_Blotter_file_path)
    Trade_Blotter=Trade_Blotter.iloc[1:]
    
    Trade_Blotter_1=Trade_Blotter[["Transaction Type",
                                  "Custodian Name",
                                  "BB Yellow Key",
                                  "Notional Quantity",
                                  "Sub Business Unit",
                                  "Strategy"]]
    
    Trade_Blotter_1["Custodian_match_name"]=Trade_Blotter_1["Custodian Name"].apply(lambda x:tuple(x.split("_")[1:3]))
    return (Trade_Blotter_1)


Trade_Blotter=Trade_Blotter()


Trade_Blotter.head(10)



SS_today=[]
Cover_today=[]


for i in range(whole.shape[0]):
    ticker=whole.iloc[i]["DW Underlying Ticker"]
    match=whole.iloc[i]["Custodian_match"]
    print(ticker,match)
    SS_qty=Trade_Blotter[(Trade_Blotter["Custodian_match_name"]==match) &
                  (Trade_Blotter["BB Yellow Key"]==ticker) &
                  (Trade_Blotter["Transaction Type"]=="SellShort")
                 ]["Notional Quantity"].sum()
    
    
    BUYCOVER_qty=Trade_Blotter[(Trade_Blotter["Custodian_match_name"]==match) &
                  (Trade_Blotter["BB Yellow Key"]==ticker) &
                  (Trade_Blotter["Transaction Type"]=="BuyToCover")
                 ]["Notional Quantity"].sum()
    
    SS_today.append(SS_qty)
    Cover_today.append(BUYCOVER_qty)
    
    
SS_net_reduce= [SS_today[i] + Cover_today[i] for i in range(len(SS_today))] 
 


whole["SS today"]=SS_today
whole["COVER today"]=Cover_today
whole["NET REDUCE"]=SS_net_reduce


whole.sort_values(by=["ticker_PM"],inplace=True)


print("negative means to reduce the PTH balance, positive means to book additional PTH into existing balance potentially - ask PM")


writer1 = pd.ExcelWriter(r"T:\Daily trades\Daily Re\PTH Adjustment\PTH_ADJUST1.xlsx", engine='xlsxwriter')

whole.to_excel(writer1, sheet_name='PTH_ADJUST',index=False)
writer1.save()

writer1.close

import win32com.client as win32
excel = win32.gencache.EnsureDispatch('Excel.Application')
wb = excel.Workbooks.Open(r"T:\Daily trades\Daily Re\PTH Adjustment\PTH_ADJUSTovernight.xlsx")
ws = wb.Worksheets("PTH_ADJUST")
ws.Columns.AutoFit()
wb.Save()
wb.Close()


#some formatting

from openpyxl import load_workbook
workbook = load_workbook(filename=r"T:\Daily trades\Daily Re\PTH Adjustment\PTH_ADJUSTovernight.xlsx")
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
ws = workbook.active

redFill = PatternFill(start_color='000000',
                   end_color='000000',
                   fill_type='solid')

for cell in ws[1]:
    cell.fill=redFill

for cell in ws["E:E"]:
    cell.number_format='#,###'
        
for cell in ws["K:K"]:
    cell.number_format='#,###'
        
for cell in ws["M:M"]:
    cell.number_format='#,###'

for cell in ws["L:L"]:
    cell.number_format='#,###'


for cell in ws[1]:
    cell.font= Font(color=colors.WHITE)



yellowFill = PatternFill(start_color='C0C0C0',
                   end_color='FFFF00',
                   fill_type='solid')

blueFill = PatternFill(start_color='FFFFFF',
                   end_color='0000FF',
                   fill_type='solid')

color_list = [blueFill,yellowFill] * 100


#ws.auto_filter.ref=ws.dimensions
#ws.auto_filter.add_filter_column(0, ["C"])
#ws.auto_filter.add_sort_condition("B2:D6")

color_pointer = 0

for cell in ws[2]:
    cell.fill=color_list[color_pointer]

for i in range(whole.shape[0]-1):
    if ws["J"+str(i+2)].value==ws["J"+str(i+3)].value:
        for cell in ws[i+3]:
            cell.fill=color_list[color_pointer]
    else:
        for cell in ws[i+3]:
            cell.fill=color_list[color_pointer+1]   
        color_pointer=color_pointer+1    
    
workbook.save(filename=r"T:\Daily trades\Daily Re\PTH Adjustment\PTH_ADJUSTovernight.xlsx")

excel_2= win32.gencache.EnsureDispatch('Excel.Application')
#wb = excel_2.Workbooks.Open(r"T:\Daily trades\Daily Re\PTH Adjustment\PTH_ADJUST.xlsx")




