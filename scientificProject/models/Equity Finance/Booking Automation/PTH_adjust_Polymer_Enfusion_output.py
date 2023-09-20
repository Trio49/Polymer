# -*- coding: utf-8 -*-
"""
Created on Oct 22 2020

@author: zzhang
"""

# Last edited by:
# Zach Zhang 06222021
# PTH update model  03222020 
import pytz
import pandas as pd
import numpy as np
# import matplotlib.pyplot as plt
from datetime import datetime, timedelta
from os import listdir
import win32com.client as win32
import pandas as pd
pd.set_option('display.max_columns', 20)
today = datetime.now(pytz.timezone('Asia/Shanghai'))
# =============================================================================
# =============================================================================
# today=today.date()
a=input("time delta:")
a=int(a)
today = today.date() - timedelta(days=a)
# 
general_date_source = today.strftime('%m/%d/%y')
today = today.strftime('%Y%m%d')

#dummy FX list
USDCNY_rate=7.2
USDTWD_rate=32
USDKRW_rate=1320

# def firm_positions():
# read the FIRM FLAT positions from some place!
# firm_positions_file_path="//paghk.local/shares/Enfusion/Polymer/"+today
firm_positions_file_path = r"P:\All\Enfusion\Polymer/"
firm_positions_list_dir = listdir(firm_positions_file_path)

firm_positions_list_dir = firm_positions_list_dir[::-1]
for file_name in firm_positions_list_dir:
    if "PTH REPORT_PTH" in file_name:
        firm_positions_file_path = firm_positions_file_path + "\\" + file_name
        print(firm_positions_file_path)
        break

####tyin
firm_positions = pd.read_csv(firm_positions_file_path)
# firm_positions = pd.read_csv("\\paghk.local\shares\Enfusion\Polymer\" + today)
#
# firm_positions_path=r"C:\Users\tyin\Documents\report_FIRM_FLAT_TYIN0322.csv"
#    firm_positions=pd.read_csv(firm_positions_path)
firm_positions = firm_positions.iloc[1:]
firm_positions.describe()
firm_positions = firm_positions[firm_positions["Book Name"].str.contains("PTH")]
firm_positions = firm_positions[firm_positions["Position"] != 0]
firm_positions = firm_positions[["Sub Business Unit",
                                 "Book Name",
                                 "Underlying BB Yellow Key",
                                 "Description",
                                 "Position",
                                 "Custodian Account Display Name",
                                 "Last Trade Date",
                                 "Deal Id",
                                 "Active Coupon Rate",
                                 "Instrument Id",
                                 'Book ID',
                                 'Custodian Acct Id']]

# to the sub-PM level to match with trade blotter strategy column
firm_positions["strats"] = firm_positions["Book Name"].apply(lambda x:x.split("(")[1].replace(")", ""))
firm_positions["Product"] = np.where(firm_positions['Custodian Account Display Name'].str.contains('ISDA'), 'swap',
                                     'cash')
#replace PCM_BORROW under Custodian Account Display Name with PCMBORROW
#
# firm_positions

firm_positions["Custodian_match"] = firm_positions["Custodian Account Display Name"].apply(
    lambda x:tuple(x.split("_")[1:3]))
    # return (firm_positions)


# firm_positions = firm_positions()
firm_positions["ticker_PM"] = firm_positions["Underlying BB Yellow Key"] + " " + firm_positions["strats"]
firm_positions["Product"] = np.where(firm_positions['Custodian Account Display Name'].str.contains('ISDA'),
                                     'swap', 'cash')
ticker_PM_list = firm_positions["ticker_PM"].tolist()
######################################################################## list all existing PM positions for those PTH names
whole = firm_positions


#

########################################################################

# read the REPORT_Polymer_Blotter_Tony_202
def Trade_Blotter():
    # Trade_Blotter_file_path=r"\\paghk.local\shares\Enfusion\Polymer\\"+today
    Trade_Blotter_file_path = r"//paghk.local/Polymer/HK/Department/All/Enfusion/Polymer/" + today
    list_dir = listdir(Trade_Blotter_file_path)

    list_dir = list_dir[::-1]

    for file_name in list_dir:
        if "Polymer_Blotter_Tony" in file_name:
            Trade_Blotter_file_path = Trade_Blotter_file_path + "\\" + file_name
            break

    Trade_Blotter = pd.read_csv(Trade_Blotter_file_path)
    Trade_Blotter = Trade_Blotter.iloc[1:]

    Trade_Blotter_1 = Trade_Blotter[["Transaction Type",
                                     "Custodian Name",
                                     "BB Yellow Key",
                                     "Notional Quantity",
                                     "Sub Business Unit",
                                     "Trade Id",
                                     "Strategy"]]

    Trade_Blotter_1["Custodian_match_name"] = Trade_Blotter_1["Custodian Name"].apply(lambda x:tuple(x.split("_")[1:3]))
    return (Trade_Blotter_1)


Trade_Blotter = Trade_Blotter()

def PTH_unwind_generation():


    net = Trade_Blotter[Trade_Blotter["Transaction Type"] == "SellShort"]

    # net["Custodian_match_name"]=net["Custodian Account Display Name"].apply(lambda x:tuple(x.split("_")[2:3]))
    net["Rate"] = 0

    # Add ID to combine custodian account display name and underlying bb yellow key
    net["ID"] = net["Custodian Name"] + net["BB Yellow Key"]
    # remove empty ID rows
    net = net[net["ID"] != ""]
    # change notional quantity to absolute value
    net["Notional Quantity"] = net["Notional Quantity"].abs()

    # filter PTH position from whole by containing string Loan/Borrow
    PTHs = whole[whole["Description"].str.contains("Loan/Borrow")]



    # combining Custodian Account Display Name and DW Underlying Ticker to create ID
    PTHs["ID2"] = PTHs["Custodian Account Display Name"] + PTHs["Underlying BB Yellow Key"]
    # Sort PTHs by Ticker_PM
    PTHs.sort_values(by=["ticker_PM"], inplace=True)
    # PTHs position changed to abs value
    PTHs["Position"] = PTHs["Position"].abs()
    # keep the PTHs columns needed
    PTHs = PTHs[["Underlying BB Yellow Key", "Description", "Position", "Custodian Account Display Name", "Deal Id",
                 "Instrument Id","Book ID","Custodian Acct Id","ID2"
                                             ]]
    # save a list of unique value of Underlying BB Yellow Key
    PTH_ticker_list = PTHs["Underlying BB Yellow Key"].unique().tolist()
    # Create a new dataframe call unwind_pth to copy net
    unwind_pth = net.copy()
    # filter unwind_pth by PTH_ticker_list
    unwind_pth = unwind_pth[unwind_pth["BB Yellow Key"].isin(PTH_ticker_list)]
    # remove empty bb yellow key
    unwind_pth = unwind_pth[unwind_pth["BB Yellow Key"] != ""]
    # remove empty custodian account display name
    unwind_pth = unwind_pth[unwind_pth["Custodian Name"] != ""]
    #unwind_pth quantity change to abs value
    unwind_pth["Notional Quantity"] = unwind_pth["Notional Quantity"].abs()
    # add new columns
    unwind_pth['deduction'] = ''
    unwind_pth['Deal Id'] = ''
    unwind_pth['Instrument Id'] = ''
    unwind_pth['Book ID'] = ''
    unwind_pth['Custodian Acct Id'] = ''
    # create a new dataframe to record unwind_pth_booking
    unwind_pth_booking = pd.DataFrame()
    #keep a column to record the original notional quantity
    unwind_pth['original_notional_quantity'] = unwind_pth['Notional Quantity']
    #reset index for PTHs and unwind_pth
    PTHs.reset_index(drop=True, inplace=True)
    unwind_pth.reset_index(drop=True, inplace=True)
    #sort PTHs by underlying bb yellow key then by position,desending
    PTHs.sort_values(by=["Underlying BB Yellow Key", "Position"], ascending=False, inplace=True)
    # loop through unwind_pth and PTHs
    for i in range(len(unwind_pth)):
        for j in range(len(PTHs)):
            # if ID and ID2 match and PTHs unwind_pth position is greater than PTHs position, the
            if unwind_pth['ID'][i] == PTHs['ID2'][j]:
                #if notional quantity is greater than PTHs position, deduction is PTHs position and notional quantity is notional quantity minus PTHs position
                #if notional quantity is less than PTHs position, deduction is notional quantity
                if unwind_pth['Notional Quantity'][i] > PTHs['Position'][j]:
                    unwind_pth['Notional Quantity'][i]=unwind_pth['Notional Quantity'][i]-PTHs['Position'][j]
                    unwind_pth['deduction'][i] = PTHs['Position'][j]
                    unwind_pth['Deal Id'][i] = PTHs['Deal Id'][j]
                    unwind_pth['Instrument Id'][i] = PTHs['Instrument Id'][j]
                    unwind_pth['Book ID'][i] = PTHs['Book ID'][j]
                    unwind_pth['Custodian Acct Id'][i] = PTHs['Custodian Acct Id'][j]
                    PTHs['Position'][j] = 0
                    #add the deduction to a new dataframe called unwind_pth_booking
                    unwind_pth_booking = unwind_pth_booking.append(unwind_pth.iloc[i])
                elif unwind_pth['Notional Quantity'][i]<=PTHs['Position'][j]:
                    unwind_pth['deduction'][i] = unwind_pth['Notional Quantity'][i]
                    unwind_pth['Deal Id'][i] = PTHs['Deal Id'][j]
                    unwind_pth['Instrument Id'][i] = PTHs['Instrument Id'][j]
                    unwind_pth['Book ID'][i] = PTHs['Book ID'][j]
                    unwind_pth['Custodian Acct Id'][i] = PTHs['Custodian Acct Id'][j]
                    unwind_pth['Notional Quantity'][i] = 0
                    # add the deduction to a new dataframe called unwind_pth_booking
                    unwind_pth_booking = unwind_pth_booking.append(unwind_pth.iloc[i])
                elif unwind_pth['Notional Quantity'][i] == 0:
                    break

                    break

    #get the string before "_" in Sub Business Unit
    unwind_pth_booking['Strat'] = unwind_pth_booking['Sub Business Unit'].apply(lambda x: x.split("_")[0])
    #if Strt has "PO" replaced it to ECMIPO
    unwind_pth_booking['Strat'] = unwind_pth_booking['Strat'].apply(lambda x: x.replace("PO", "ECMIPO"))
    #
    #get the custodian name from the string between the two "_" in Custodian_match_name
    # #for rows with Strat not having PCM, use the string after the first "_" in Custodian_match_name
    # for i in range(len(unwind_pth_booking)):
    #     #if the Custodian Name does not have string" PCM" in it
    #     if " PCM" not in unwind_pth_booking['Custodian Name'][i]:
    #
    unwind_pth_booking['Custodian'] = unwind_pth_booking['Custodian_match_name'].apply(lambda x: x[1])
    #     else:#use the string between the third "_" and the fourth "_" in Custodian Name
    #         unwind_pth_booking['Custodian'] = unwind_pth_booking['Custodian Name'].apply(lambda x: x[3])

    #keep the columns needed
    unwind_pth_booking = unwind_pth_booking[['Strat',
                                             'Custodian',
                                             'Custodian Name',
                                             'BB Yellow Key',
                                             'deduction',
                                             'Rate',
                                             'Deal Id',
                                             'Instrument Id',
                                             'Book ID',
                                             'Custodian Acct Id'

                                             ]]
    #if Strat = PCM and Custodian=BORROW, change the Start to PCM_BORROW then change the Custodian to the string before "_ISDA" in Custodian Name
    unwind_pth_booking.loc[(unwind_pth_booking['Strat'] == 'PCM') & (unwind_pth_booking['Custodian'] == 'BORROW'), 'Strat'] = 'PCM_BORROW'
    #reset index
    unwind_pth_booking.reset_index(drop=True, inplace=True)
    #get the string between third "_" and fourth "_" in Custodian Name
    unwind_pth_booking['Custodian2'] = unwind_pth_booking['Custodian Name'].apply(lambda x: x.split("_")[3])
    unwind_pth_booking['Custodian']=np.where (unwind_pth_booking['Custodian'] == 'BORROW', unwind_pth_booking['Custodian2'], unwind_pth_booking['Custodian'])
    unwind_pth_booking = unwind_pth_booking[['Strat',
                                                 'Custodian',

                                                 'BB Yellow Key',
                                                 'deduction',
                                                 'Rate',
                                                 'Deal Id',
                                                 'Instrument Id',
                                                 'Book ID',
                                                 'Custodian Acct Id'

                                                 ]]
    return unwind_pth_booking,unwind_pth
unwind_pth_booking,unwind_pth = PTH_unwind_generation()

unwind_pth_booking.to_csv(r'P:\Operations\Polymer - Middle Office\PTH and FX booking\PTH unwind/PTH unwind'+today+'.csv', index=False)

##############create a new dataframe called unwind_pth_booking_shell to record the unwind_pth_booking




# def unwind_PTH_trade():
Unwind_trades_X=unwind_pth_booking.copy()
Unwind_trades_X_email=Unwind_trades_X[['Strat','Custodian','BB Yellow Key','deduction','Deal Id']]
Unwind_trades_X_email.to_excel(r'P:\Operations\Polymer - Middle Office\PTH and FX booking\PTH unwind/PTHemail.xlsx', index=False)
def formatting_enfusion_input():
# convert the unwind_trades_deduction into an Enfusion consumable format

# example
    example = pd.read_excel(r'P:\Operations\Polymer - Middle Office\PTH and FX booking/PTH Unwind_sample.xlsx',
                            sheet_name='Sheet1')
    # save the columns of the example into a list
    example_list = example.columns.tolist()
    # assign the list to a new dataframe called unwind_trades_shell
    Unwind_trades_shell = pd.DataFrame(columns=example_list).reset_index(drop=True)
    # for each row in Unwind_trades_deduction, append the row to Unwind_trades_shell
    for i in range(len(Unwind_trades_X)):
        Unwind_trades_shell = Unwind_trades_shell.append(Unwind_trades_X.iloc[i])

    # enrich some the columns with fixed values
    Unwind_trades_shell['Repo Subtype'] = 'StockLoanBorrow'
    Unwind_trades_shell['Transaction Type'] = 'Sell'
    Unwind_trades_shell['Booing Status'] = 'Complete'
    Unwind_trades_shell['Haircut Rate'] = 1
    Unwind_trades_shell['Settle CCY'] = 'USD'
    Unwind_trades_shell['Reference Instrument Identifier Type'] = 'BLOOMBERG YELLOW'
    Unwind_trades_shell['Repo Type'] = 'Open'
    Unwind_trades_shell['Price Resets Daily'] = 'TRUE'
    Unwind_trades_shell['Accrued Override'] = 0
    Unwind_trades_shell['Interest Settles On Unwind Date'] = 'TRUE'
    Unwind_trades_shell['Other Borrow Fee'] = 0
    Unwind_trades_shell['Payment Type'] = 'FixedRate'
    Unwind_trades_shell['Payment Frequency'] = 'Monthly'
    Unwind_trades_shell['Day Count'] = 'ACTUAL_360'
    Unwind_trades_shell['Business Day Convention'] = 'FOLLOWING'
    Unwind_trades_shell['Compounding Frequency'] = 'Simple'
    Unwind_trades_shell['Booking Status'] = 'Complete'
    Unwind_trades_shell['Clean Price'] = 0
    # #read the account mapping table
    # account_mapping = pd.read_excel(r'P:\Operations\Polymer - Middle Office\PTH and FX booking\Custodian_id_mapping.xlsx',sheet_name='Sheet1')
    # #rename the acocunt_mapping columns
    # #remove Account Id and Counterparty Code columns from Unwind_trades_shell
    Unwind_trades_shell = Unwind_trades_shell.drop(['Strat'], axis=1)
    #replace Account Id with Custodian Acct Id
    #replace Book Id with Book ID
    Unwind_trades_shell=Unwind_trades_shell.drop(['Account Id'],axis=1)
    Unwind_trades_shell = Unwind_trades_shell.rename(columns={'Custodian Acct Id': 'Account Id'})
    Unwind_trades_shell=Unwind_trades_shell.drop(['Book Id'],axis=1)
    Unwind_trades_shell = Unwind_trades_shell.rename(columns={'Book ID': 'Book Id'})
    Unwind_trades_shell=Unwind_trades_shell.drop(['Booing Status'],axis=1)
    Unwind_trades_shell=Unwind_trades_shell.drop(['Rate'],axis=1)
    # #enrich the Unwind_trades_shell with the account id based on the custodian account display name
    # Unwind_trades_shell = pd.merge(Unwind_trades_shell, account_mapping, on='Custodian Account Display Name', how='left')
    # Trade Date and Settle Date is today_text
    Unwind_trades_shell['Trade Date'] = general_date_source
    Unwind_trades_shell['Settle Date'] = general_date_source

    # CNY rate is (1,1) in the USDCNY dataframe

    # Trade/Settle FX Rate is USDCNY
    Unwind_trades_shell['Trade/Settle FX Rate'] = ''

    # remove 'Counterparty Code' and rename Custodian to Counterparty Code
    Unwind_trades_shell = Unwind_trades_shell.drop(['Counterparty Code'], axis=1)
    Unwind_trades_shell = Unwind_trades_shell.rename(columns={'Custodian': 'Counterparty Code'})
    #remoe Enfusion Instrument Id and rename Instrument Id to Enfusion Instrument Id
    Unwind_trades_shell = Unwind_trades_shell.drop(['Enfusion Instrument Id'], axis=1)
    Unwind_trades_shell = Unwind_trades_shell.rename(columns={'Instrument Id': 'Enfusion Instrument Id'})
    #remove BB yellow key and rename BB Yellow Key to Reference Instrument Identifier
    Unwind_trades_shell = Unwind_trades_shell.drop(['Reference Instrument Identifier'], axis=1)
    Unwind_trades_shell = Unwind_trades_shell.rename(columns={'BB Yellow Key': 'Reference Instrument Identifier'})
    #remove Notional Quantity and rename deduction to Notional Quantity
    Unwind_trades_shell = Unwind_trades_shell.drop(['Notional Quantity'], axis=1)
    Unwind_trades_shell = Unwind_trades_shell.rename(columns={'deduction': 'Notional Quantity'})
    # fill the nan values in the Unwind_trades_shell with ''
    Unwind_trades_shell = Unwind_trades_shell.fillna('')

    Unwind_trades_shell['Price Resets Daily'] = Unwind_trades_shell['Price Resets Daily'].astype(bool)
    Unwind_trades_shell['Interest Settles On Unwind Date'] = Unwind_trades_shell[
        'Interest Settles On Unwind Date'].astype(bool)
    # add a replacement dictionary for Counterparty Code
    replace_dict = {'BAML':'MLCO', 'GS':'GSCO', 'jpm':'JPM', 'UBS':'UBS','MS':'MSCO'}
    # replace the Counterparty Code with the replacement dictionary
    Unwind_trades_shell['Counterparty Code'] = Unwind_trades_shell['Counterparty Code'].replace(replace_dict)
    # Unwind_trades_shell = Unwind_trades_shell.astype(str).apply(lambda x: x.str.replace('.', ''))
    #remove the 0 Notional Quantity
    Unwind_trades_shell=Unwind_trades_shell[Unwind_trades_shell['Notional Quantity']!=0]
    # reset the index
    Unwind_trades_shell = Unwind_trades_shell.reset_index(drop=True)
    # remove 1st column
    # Unwind_trades_shell = Unwind_trades_shell.drop(Unwind_trades_shell.columns[0], axis=1)
    # save the dataframe as an excel
    return Unwind_trades_shell
Unwind_trades_shell=formatting_enfusion_input()

Unwind_trades_shell.to_excel(
    r"P:\Operations\Polymer - Middle Office\PTH and FX booking/PTH unwind/PTH_Upload_test" + today + '.xlsx', index=False)
#     return Unwind_trades_shell
from openpyxl import load_workbook

wb = load_workbook(r"P:\Operations\Polymer - Middle Office\PTH and FX booking/PTH unwind/PTH_Upload_test" + today + '.xlsx')
ws = wb.active

for row in ws.iter_rows(min_row=2, min_col=2, max_col=3):
    for cell in row:
        if cell.value is not None:
            date_object = datetime.strptime(cell.value, "%m/%d/%y")
            general_date = date_object.toordinal() - datetime(1900, 1, 1).toordinal() + 2
            cell.value = general_date

wb.save(r"P:\Operations\Polymer - Middle Office\PTH and FX booking/PTH unwind/PTH_Upload_test" + today + '.xlsx')

#send a Unwind_trade_X to the team in an email
def send_pth():

    outlook = win32.Dispatch('Outlook.Application')
    mail = outlook.CreateItem(0)

    mail.to = 'polymerops@polymercapital.com;tyin@polymercapital.com;douglasl@polymercapital.com'

    sub = "PTH Return and Utilization" + "  " + today

    mail.Subject = sub
    mail.Body = 'Message body testing'
    with pd.ExcelWriter(r'P:\Operations\Polymer - Middle Office\PTH and FX booking\PTH unwind/PTHemail.xlsx',engine='xlsxwriter') as writer:
        Unwind_trades_X_email.to_excel(writer, sheet_name='Sheet1')
        html= Unwind_trades_X_email.to_html()
        text_file = open(r"P:\Operations\Polymer - Middle Office\PTH and FX booking\pthunwind.htm",'w')
        text_file.write(html)
        text_file.close()
    # HtmlFile = open(r'C:\Users\tyin\Documents\\'+newFileName+'.htm', 'r')
    # HtmlFile = open(r"P:\Operations\Polymer - Middle Office\PTH and FX booking\pthunwind.htm",'w')
    # source_code = HtmlFile.read()
    # HtmlFile.close()
    # html=wb.to_html(classes='table table-striped')
    # html=firm_positions.to_html(classes='table table-striped')
    # mail.HTMLBody = source_code
    mail.HTMLBody = html
    mail.Send()
    print('done')
send_pth()