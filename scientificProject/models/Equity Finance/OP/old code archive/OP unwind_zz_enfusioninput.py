# -*- coding: utf-8 -*-
"""
Created on Thu Mar  9 14:40:40 2023

@author: zzhang
"""
# import numpy as np
# import pdblp
from datetime import datetime, timedelta
from os import listdir
import blpapi
import pandas as pd
import pytz
from xbbg import blp

a=input('Time Delta')
#convert a to int
a=int(a)
today = datetime.now(pytz.timezone('Asia/Shanghai'))
today = today.date() - timedelta(days=a)
today_text = today.strftime('%Y%m%d')


#save a toordinal version of today_text
general_date_source=today.strftime('%m/%d/%y')
# general_date_object=datetime.strptime(general_date_source,'%m/%d/%y')
# general_date=general_date_object.toordinal()-datetime(1900, 1, 1).toordinal()+2
today_datetime=today
############remove when running officially 
# today=today.strftime('%Y%m%d')
# today=today.date()-timedelta(days=1)
# today=today.date()
today = today.strftime('%Y%m%d')

# all_PM_expiryblotter_file_path=r"\\paghk.local\shares\Enfusion\Polymer\\"+today
all_PM_expiryblotter_file_path = r"//paghk.local/Polymer/HK/Department/All/Enfusion/Polymer/" + today

firm_positions_list_dir = listdir(all_PM_expiryblotter_file_path)

firm_positions_list_dir1 = firm_positions_list_dir[::-1]
# define the op position path and report name
for file_name in firm_positions_list_dir1:

    if "Polymer_OP" in file_name:
        print(file_name)

        all_PM_expiryblotter_file_path = all_PM_expiryblotter_file_path + "\\" + file_name
        break
# read OP position file
OP_pos = pd.read_csv(all_PM_expiryblotter_file_path, engine='python', thousands=',')
OP_pos['PM'] = OP_pos['Book Name'].str[-5:]
replace_dict = {'BAML':'baml', 'Goldman Sachs':'gs', 'JP Morgan':'jpm', 'UBS':'ubs'}
OP_pos['Prime'] = OP_pos['Prime'].replace(replace_dict, regex=True)
OP_pos['ID2'] = OP_pos['PM'] + OP_pos['Prime'] + OP_pos['Underlying BB Yellow Key']
#keep the columns we need
OP_pos = OP_pos[
    ['ID2', 'Underlying BB Yellow Key', 'Position', 'Custodian Account Display Name', 'Deal Id', 'Instrument Id','Book ID','Custodian Acct Id']]
def unwind_OP_trades_generation(OP_pos):
    # read unwind trade
    Unwind_trades = pd.read_csv(
        '//paghk.local/Polymer/Department/All/FinanceTradingMO/outperformance swap/Daily booking/' + today_text + '_unwind_OP_trades.csv',
        engine='python', thousands=',')
    Unwind_trades['BBcode'] = Unwind_trades['BBG Code'] + ' ' + 'Equity'
    Unwind_trades['Internal Account'] = Unwind_trades['Internal Account'].str[-5:]
    Unwind_trades['ID'] = Unwind_trades['Internal Account'] + Unwind_trades['Broker'] + Unwind_trades['BBcode']
    # check = input('is the PTH file latest? yes/no: ')
    Unwind_trades = Unwind_trades[['Internal Account', 'BBcode', 'Broker', 'unwind_op_quantity', 'ID']]
    # remove trades with 0 unwind_op_quantity
    Unwind_trades = Unwind_trades[Unwind_trades['unwind_op_quantity'] != 0]
    # sum the unwind_op_quantity with the same ID, Internal Account, Broker, BBcode
    Unwind_trades_X = Unwind_trades.groupby(['ID', 'Internal Account', 'Broker', 'BBcode']).sum()
    Unwind_trades_X = Unwind_trades_X.reset_index()
    # cross check Unwind_trades_X ID with OP_pos ID2,if there is any ID matched, for each matched ID, minus the unwind_op_quantity from the position and record each value deducted
    Unwind_trades_X['Unwind_op_quantity_total'] = Unwind_trades_X['unwind_op_quantity']
    # Create a new dataframe to store the deduction row with deduction quantity,internal account, broker, BBcode, deal id and instrument id
    Unwind_trades_deduction = pd.DataFrame(
        columns=['Internal Account', 'Broker', 'BBcode', 'deduction', 'Custodian Account Display Name', 'Deal Id',
                 'Instrument Id'])
    # create a deduction column under Unwind_trades_X
    Unwind_trades_X['deduction'] = 0
    # create a deal id column under Unwind_trades_X
    Unwind_trades_X['Deal Id'] = ''
    # create a instrument id column under Unwind_trades_X
    Unwind_trades_X['Instrument Id'] = ''
    # create a Book ID and Custodian Acct Id column under Unwind_trades_X
    Unwind_trades_X['Book Id'] = ''
    Unwind_trades_X['Account Id'] = ''
    # create a column to store Custodian Account Display Name
    Unwind_trades_X['Custodian Account Display Name'] = ''
    for i in range(len(Unwind_trades_X)):
        for j in range(len(OP_pos)):
            # if the ID in Unwind_trades_X matches the ID2 in OP_pos, and the unwind_op_quantity is greater than the position, then minus the position from the unwind_op_quantity and record the value deducted
            if Unwind_trades_X['ID'][i] == OP_pos['ID2'][j]:
                # if the unwind_op_quantity is greater than the position, then minus the position from the unwind_op_quantity and record the value deducted
                # append the deal id and instrument id to the Unwind_trades_X and save the row to a new dataframe

                if Unwind_trades_X['unwind_op_quantity'][i] >= OP_pos['Position'][j]:
                    Unwind_trades_X['unwind_op_quantity'][i] = Unwind_trades_X['unwind_op_quantity'][i] - \
                                                               OP_pos['Position'][j]
                    Unwind_trades_X['deduction'][i] = OP_pos['Position'][j]
                    # inherit custodian account display name, deal id and instrument id from OP_pos
                    Unwind_trades_X['Deal Id'][i] = OP_pos['Deal Id'][j]
                    Unwind_trades_X['Book Id'][i] = OP_pos['Book ID'][j]
                    Unwind_trades_X['Account Id'][i] = OP_pos['Custodian Acct Id'][j]
                    Unwind_trades_X['Instrument Id'][i] = OP_pos['Instrument Id'][j]
                    Unwind_trades_X['Custodian Account Display Name'][i] = OP_pos['Custodian Account Display Name'][j]
                    # append ['deduction', 'Internal Account', 'Broker', 'BBcode', 'Deal Id','Instrument Id'] to the Unwind_trades_deduction
                    Unwind_trades_deduction = Unwind_trades_deduction.append(Unwind_trades_X.iloc[i])


                elif Unwind_trades_X['unwind_op_quantity'][i] < OP_pos['Position'][j]:
                    # if the unwind_op_quantity is less than the position, then minus the unwind_op_quantity from the position and record the value deducted
                    Unwind_trades_X['deduction'][i] = Unwind_trades_X['unwind_op_quantity'][i]
                    Unwind_trades_X['Deal Id'][i] = OP_pos['Deal Id'][j]
                    Unwind_trades_X['Book Id'][i] = OP_pos['Book ID'][j]
                    Unwind_trades_X['Account Id'][i] = OP_pos['Custodian Acct Id'][j]
                    Unwind_trades_X['Instrument Id'][i] = OP_pos['Instrument Id'][j]
                    Unwind_trades_X['Custodian Account Display Name'][i] = OP_pos['Custodian Account Display Name'][j]

                    # append ['deduction', 'Internal Account', 'Broker', 'BBcode', 'Deal Id','Instrument Id'] to the Unwind_trades_deduction
                    Unwind_trades_deduction = Unwind_trades_deduction.append(Unwind_trades_X.iloc[i])





                break


    #remove the rows with 0 deduction from Unwind_trades_deduction
    Unwind_trades_deduction = Unwind_trades_deduction[Unwind_trades_deduction['deduction'] != 0]
    #save the file
    Unwind_trades_deduction.to_csv(r"P:\All\FinanceTradingMO\outperformance swap\Daily booking\unwind_new" + today +'.csv')

    #convert the unwind_trades_deduction into an Enfusion consumable format

    #example
    example=pd.read_excel(r'P:\Operations\Polymer - Middle Office\PTH and FX booking/OP_Unwind.xlsx',sheet_name='Sheet1')
    #save the columns of the example into a list
    example_list = example.columns.tolist()
    #assign the list to a new dataframe called unwind_trades_shell
    Unwind_trades_shell = pd.DataFrame(columns=example_list).reset_index(drop=True)
    #for each row in Unwind_trades_deduction, append the row to Unwind_trades_shell
    for i in range(len(Unwind_trades_deduction)):
        Unwind_trades_shell = Unwind_trades_shell.append(Unwind_trades_deduction.iloc[i])

    # enrich some the columns with fixed values
    Unwind_trades_shell['Repo Subtype'] = 'StockLoanBorrow'
    Unwind_trades_shell['Transaction Type'] = 'BuyToCover'
    Unwind_trades_shell['Booing Status'] = 'Complete'
    Unwind_trades_shell['Haircut Rate'] = 1
    Unwind_trades_shell['Settle CCY'] = 'USD'
    Unwind_trades_shell['Reference Instrument Identifier Type'] = 'BLOOMBERG YELLOW'
    Unwind_trades_shell['Repo Type'] = 'Open'
    Unwind_trades_shell['Price Resets Daily'] = 'TRUE'
    Unwind_trades_shell['Accrued Override'] = 0
    Unwind_trades_shell['Interest Settles On Unwind Date'] = 'TRUE'
    Unwind_trades_shell['Other Borrow Fee'] = 0
    Unwind_trades_shell['Payment Type'] = 'FixedRate'
    Unwind_trades_shell['Payment Frequency'] = 'Monthly'
    Unwind_trades_shell['Day Count'] = 'ACTUAL_360'
    Unwind_trades_shell['Business Day Convention'] = 'FOLLOWING'
    Unwind_trades_shell['Compounding Frequency'] = 'Simple'
    Unwind_trades_shell['Booking Status'] = 'Complete'
    # #read the account mapping table
    # account_mapping = pd.read_excel(r'P:\Operations\Polymer - Middle Office\PTH and FX booking\Custodian_id_mapping.xlsx',sheet_name='Sheet1')
    # #rename the acocunt_mapping columns
    # account_mapping.columns = ['Custodian Account Display Name', 'Counterparty Code','Account Id']
    # #remove Account Id and Counterparty Code columns from Unwind_trades_shell
    # Unwind_trades_shell = Unwind_trades_shell.drop(['Account Id', 'Counterparty Code'], axis=1)
    # #enrich the Unwind_trades_shell with the account id based on the custodian account display name
    # Unwind_trades_shell = pd.merge(Unwind_trades_shell, account_mapping, on='Custodian Account Display Name', how='left')
    #Trade Date and Settle Date is today_text
    Unwind_trades_shell['Trade Date'] = general_date_source
    Unwind_trades_shell['Settle Date'] = general_date_source
    # #conver the Trade Date and Settle Date to Int
    # Unwind_trades_shell['Trade Date'] = Unwind_trades_shell['Trade Date'].astype(int)
    # Unwind_trades_shell['Settle Date'] = Unwind_trades_shell['Settle Date'].astype(int)
    #call USDCNY currency rate from bloomberg
    USDCNY = blp.bdh('USDCNY Curncy', 'PX_LAST', today_text, today_text)
    #the USDCNY rate is (1,1) in the USDCNY dataframe
    #If USDCNY rate is not available, use the previous manual input FX
    if USDCNY.empty:
        USDCNY_rate = 7.1   #use the previous manual input FX
        USDCNY_rate_text=str(USDCNY_rate)
        print(USDCNY_rate_text+' is used as the USDCNY rate')
    else:
     USDCNY_rate = USDCNY.iloc[0,0]
    USDCNY_rate = float(USDCNY_rate)
    #Trade/Settle FX Rate is USDCNY
    Unwind_trades_shell['Trade/Settle FX Rate'] = USDCNY_rate

    #remove 'Notional Quantity' column
    Unwind_trades_shell = Unwind_trades_shell.drop(['Notional Quantity'], axis=1)
    #rename deduction to Notional Quantity
    Unwind_trades_shell = Unwind_trades_shell.rename(columns={'deduction': 'Notional Quantity'})
    #remove original reference instrument identifier column
    Unwind_trades_shell = Unwind_trades_shell.drop(['Reference Instrument Identifier'], axis=1)
    #rename BBcode to Reference Instrument Identifier
    Unwind_trades_shell = Unwind_trades_shell.rename(columns={'BBcode': 'Reference Instrument Identifier'})
    #rename Broker to Counterparty Code
    Unwind_trades_shell = Unwind_trades_shell.rename(columns={'Broker': 'Counterparty Code'})
    #rename Instrument Id to Enfusion Instrument Id
    Unwind_trades_shell = Unwind_trades_shell.rename(columns={'Instrument Id': 'Enfusion Instrument Id'})
    #fill the nan values in the Unwind_trades_shell with ''
    Unwind_trades_shell = Unwind_trades_shell.fillna('')
    #remove columns ID,unwind_op_quantity, unwind_op_quantity,Unwind_op_quantity_total and Booing Status
    Unwind_trades_shell = Unwind_trades_shell.drop(['ID','Internal Account','Custodian Account Display Name','unwind_op_quantity', 'unwind_op_quantity', 'Unwind_op_quantity_total','Booing Status'], axis=1)
    #Convert Price Resets Daily and Interest Settle On Unwind Date to BOOLEAN
    Unwind_trades_shell['Price Resets Daily'] = Unwind_trades_shell['Price Resets Daily'].astype(bool)
    Unwind_trades_shell['Interest Settles On Unwind Date'] = Unwind_trades_shell['Interest Settles On Unwind Date'].astype(bool)
    #add a replacement dictionary for Counterparty Code
    replace_dict = {'baml':'MLCO', 'gs':'GSCO', 'jpm':'JPM', 'ubs':'UBS'}
    #replace the Counterparty Code with the replacement dictionary
    Unwind_trades_shell['Counterparty Code'] = Unwind_trades_shell['Counterparty Code'].replace(replace_dict)
    # Unwind_trades_shell = Unwind_trades_shell.astype(str).apply(lambda x: x.str.replace('.', ''))
    #reset the index
    Unwind_trades_shell = Unwind_trades_shell.reset_index(drop=True)
    #remove 1st column
    # Unwind_trades_shell = Unwind_trades_shell.drop(Unwind_trades_shell.columns[0], axis=1)
    #save the dataframe as an excel
    Unwind_trades_shell.to_excel(r"P:\Operations\Polymer - Middle Office\PTH and FX booking/OP_Upload_test" + today +'.xlsx', index=False)
unwind_OP_trades_generation(OP_pos)
#define another function to convert the opbooking file to trades_shell format and save it to OP_Upload_test
def book_OP_trades_generation(OP_pos):
    #read the csv
    OP_new_trades = pd.read_csv(
        '//paghk.local/Polymer/Department/All/FinanceTradingMO/outperformance swap/Daily booking/' + today_text + '_OP_booking_file.csv',
        engine='python', thousands=',')

from openpyxl import load_workbook

wb = load_workbook(r"P:\Operations\Polymer - Middle Office\PTH and FX booking/OP_Upload_test" + today +'.xlsx')
ws = wb.active

for row in ws.iter_rows(min_row=2, min_col=2, max_col=3):
    for cell in row:
        if cell.value is not None:
            date_object = datetime.strptime(cell.value, "%m/%d/%y")
            general_date = date_object.toordinal() - datetime(1900, 1, 1).toordinal()+2
            cell.value = general_date

wb.save(r"P:\Operations\Polymer - Middle Office\PTH and FX booking/OP_Upload_test" + today +'.xlsx')
print('done')