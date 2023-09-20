# -*- coding: utf-8 -*-
"""
Created on Wed Mar 10 10:03:33 2021

@author: zzhang
"""


import pytz
import pandas as pd
import numpy as np
import openpyxl
#import matplotlib.pyplot as plt 
from datetime import datetime, timedelta
from os import listdir
from openpyxl import load_workbook
import win32com.client as win32
from IPython.display import HTML

today=datetime.now(pytz.timezone('Asia/Shanghai'))

############remove when running officially 
today=today.date()
# today=today.date()-timedelta(days=1)
#today=today.date()
today=today.strftime('%Y%m%d')

# all_PM_expiryblotter_file_path=r"\\paghk.local\shares\Enfusion\Polymer\\"+today
all_PM_expiryblotter_file_path=r"P:\All\Enfusion\Polymer/"+today
firm_positions_list_dir=listdir(all_PM_expiryblotter_file_path)

firm_positions_list_dir1 = firm_positions_list_dir[::-1]

for file_name in firm_positions_list_dir1:
    
    
    if "Polymer_Corporate Action trades" in file_name:
        print(file_name)

        all_PM_expiryblotter_file_path=all_PM_expiryblotter_file_path+"\\"+file_name
        break

#Trade_Blotter_Data_allPM=pd.read_csv(all_PM_tradeblotter_file_path)
#load EOD blotter
CAreport=pd.read_csv(all_PM_expiryblotter_file_path,engine='python')
CAreport.to_excel(r"T:\Daily trades\Daily Re\Python_Trade_Blotter\CAreport.xlsx",index=False)
#open loader
# wb = openpyxl.load_workbook('Expiry.xlsx')
# ws = wb.active
# rngs = list(ws.values)
# d = {}
# for row in rngs[1:]:
#     if row[9] in d.keys():
#         d[row[9]] += [row]
#     else:
#         d[row[9]] = [row]

# nwb = openpyxl.Workbook()
# # key is PM , Position is Item
# for pm, position in sorted(d.items()):
#     nws=nwb.create_sheet(pm)  # create workbook
#     nws.append(rngs[0])
#     for expiry in position:
#         nws.append(expiry)
# nwb.remove(nwb.worksheets[0])
# nwb.save(r"T:\Daily trades\Daily Re\Python_Trade_Blotter\Expiry report\Expiry per pm.xlsx")
pmlist=CAreport['FundShortName'].unique
pd.options.display.float_format = '{:,.2f}'.format

with pd.ExcelWriter(r"T:\Daily trades\Daily Re\Python_Trade_Blotter\CAreport.xlsx",engine='xlsxwriter') as writer:
  for n in pmlist():
   CAreport[CAreport['FundShortName']==n].to_excel(writer,sheet_name=n,index=None)

  for n in pmlist():
    html = CAreport[CAreport['FundShortName']==n].to_html(index=False)
    text_file = open("T:\Daily trades\Daily Re\Python_Trade_Blotter\Expiry report\CA.htm",'w')
    text_file.write(html)
    text_file.close()

    outlook = win32.Dispatch('Outlook.Application')
    mail = outlook.CreateItem(0)
      
    mail.to = 'polymerops@polymercapital.com'
        
    sub = "Corp Act " + today + " (" + n + ")"
        
        
    mail.Subject = sub
    mail.Body ='Message body testing'
        
        #HtmlFile = open(r'C:\Users\tyin\Documents\\'+newFileName+'.htm', 'r')
    HtmlFile=open(r'T:\Daily trades\Daily Re\Python_Trade_Blotter\Expiry report\CA.htm', 'r')
    source_code = HtmlFile.read() 
    HtmlFile.close()
        #html=wb.to_html(classes='table table-striped')
        #html=firm_positions.to_html(classes='table table-striped')
    mail.HTMLBody= source_code
    mail.Send() 