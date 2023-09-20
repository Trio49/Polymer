import os
import pytz
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
pd.set_option('mode.chained_assignment', None)
from pandas.tseries.offsets import BDay
from os import listdir
import math

#define datetime
x=input('time delta(0 for today, 1 for yesterday, 2 for 2 days ago, etc):')
x=int(x)
today=datetime.now(pytz.timezone('Asia/Shanghai'))
print(today)
today=today.date()-timedelta(days=x)
print(today)
yesterday=today+timedelta(days=-1)
print(yesterday)
yesterday=today-BDay(1)
print(yesterday)

today_str=today.strftime('%Y%m%d')
today_str_name=today.strftime('%Y-%m-%d')
yesterday_str=yesterday.strftime('%Y%m%d')
yesterday_str_name=yesterday.strftime('%Y-%m-%d')
DAY_EOD_path=r"P:\All\Enfusion\Polymer\\"+today_str
MAPPING_path=r"P:\Operations\Polymer - Middle Office\EOD Check\PM_MAPPING.xlsx"

#READING File 
DAY_EOD_path_dir=listdir(DAY_EOD_path)
DAY_EOD_path_dir = DAY_EOD_path_dir[::-1]


for DAY_EOD_file in DAY_EOD_path_dir:
    if "Trade Blotter - PM ALL Overnight (New)" in DAY_EOD_file:
        print(DAY_EOD_file)
        DAY_EOD_path = DAY_EOD_path + "\\" + DAY_EOD_file
        break
    
# hard code for checking 
#DAY_EOD_path = "P:\All\Enfusion\Polymer\\20230824\REPORT_Polymer_Trade Blotter - PM ALL (New)_20230824183020.csv"
#yesterday_str="2023-08-21"

print(DAY_EOD_path)

# extract both pm_mappaing data, trader list, EOD booking data, mainbook map, Subusiness map (for hierarchy)
DAY_EOD_File=pd.read_csv(DAY_EOD_path)
PM_MAP_File=pd.read_excel(MAPPING_path, sheet_name="PM_MAPPING")
TRADER_Map=pd.read_excel(MAPPING_path, sheet_name="Trader_List")
MAINBOOK_MAP=pd.read_excel(MAPPING_path, sheet_name="MAINBOOK")
SBU_MAP=pd.read_excel(MAPPING_path, sheet_name="Sub Bus Unit Mapping")
NON_OVERNIGHT_MAP=pd.read_excel(MAPPING_path, sheet_name="Non-overnight Country")


#Part 0 
# Country_list = DAY_EOD_File['Exchange Country Code'].tolist()
# abc=np.unique(Country_list)
Country_LIST = NON_OVERNIGHT_MAP['Non-overnight'].tolist()

def OVERNIGHT_CHECK(row):
    Country = row['Exchange Country Code']
    
    if Country in  Country_LIST: 
        return "NON OVERNIGHT"
    else:
        return "GOOD"
    
DAY_EOD_File['OVERNIGHT CHECKS'] = DAY_EOD_File.apply(OVERNIGHT_CHECK, axis=1)


OVERNIGHT_DAY = DAY_EOD_File[['OVERNIGHT CHECKS']]

#For Checking purposes, Check for any false value
column_counts_Overnight = DAY_EOD_File['OVERNIGHT CHECKS'].value_counts()

False_Check_OVERNIGHT= DAY_EOD_File[DAY_EOD_File['OVERNIGHT CHECKS'] == 'NON OVERNIGHT']

False_Check_OVERNIGHT_Final = False_Check_OVERNIGHT[["Counterparty","Custodian Name","Exchange Country Code","BB Yellow Key","Trade Id","OVERNIGHT CHECKS"]]

# print success message if no false result for users 
# if False_Check_OVERNIGHT_Final.shape[0]==0:
#     print ('Overnight Check Pass!') 
# else:
#     print ('Please Check for Overnight')



# Part 1: HIERACHY CHECKS 
# 3 cases; TORA-STAGED, TORA-DROPCOPY, Manual

# Convert DataFrame PM_MAPPING to a dictionary-like mapping
lookup_dict = PM_MAP_File.set_index('HIERARCHY ')['PM LOGIN'].to_dict()
lookup_dict_AF = PM_MAP_File.set_index('HIERARCHY ')['FUND NAME'].to_dict()

# Create a List of trader LOGIN values from TRADER_Map
trader_login_values = TRADER_Map['PM LOGIN'].tolist()
trader_pm = TRADER_Map['PM'].tolist()

# Map the PM Login with sub business unit
def HIERARCHY_CHECK_TEST(row):
    TORA = row['PAGStagedVsDropcopy']
    
    if TORA == 'Manual': 
        return "~chk manually~"
    elif TORA == "TORA-STAGED":
        return lookup_dict.get(row['Sub Business Unit'], "Unknown")
    elif TORA == "TORA-DROPCOPY":
        return row['PolymerToraInternalAccount']
    else:
        return "ISSUE AHHHHHH"

DAY_EOD_File['HIERARCHY CHECKS'] = DAY_EOD_File.apply(HIERARCHY_CHECK_TEST, axis=1)

# Function to check whether pm login email match with the order creation user (TORA-STAGED) or 
#[TORA-DROPCOPY] match with Business_Unit/ match with sub_business_unit if it is in SBU_list (aw_alpha, etc....)

SBU_LIST = SBU_MAP['SBU'].tolist()

def HIERARCHY_T_F(row):
    TORA = row['PAGStagedVsDropcopy']
    HIERARCHY = row['HIERARCHY CHECKS']
    Creation_User = row['Order Creation User']
    Business_Unit = row['Business Unit']
    SUB_Business_Unit = row['Sub Business Unit']
    TEMP = TRADER_Map[TRADER_Map["PM LOGIN"]==Creation_User]
    TEMP_list = TEMP['PM'].tolist()
    
    if TORA == 'Manual': 
        return "~chk manually~"
    
    elif TORA == "TORA-DROPCOPY" and HIERARCHY == Business_Unit or HIERARCHY in SBU_LIST:
        return "TRUE"
    else:
        if TORA == "TORA-STAGED" and HIERARCHY == Creation_User:
            return "TRUE"
        elif TORA == "TORA-STAGED" and Creation_User in trader_login_values and SUB_Business_Unit in TEMP_list: #If both not match, check whether the order creation users fall in any excution trader list
            return "TRUE"
        elif TORA == "TORA-STAGED" and Creation_User in trader_login_values and SUB_Business_Unit not in TEMP_list: #If both not match, check whether the order creation users fall in any excution trader list
            return "Match for trader only!"
        else:
            return "FALSE"

DAY_EOD_File['TRUE/FALSE'] = DAY_EOD_File.apply(HIERARCHY_T_F, axis=1)


# Select the final result column into a separate dataframe
HIERARCHY_DAY_PM = DAY_EOD_File[[ 'HIERARCHY CHECKS' , 'TRUE/FALSE']]

#For Checking purposes, Check for any false value (include manual)
column_counts_HIERARCHY = HIERARCHY_DAY_PM['TRUE/FALSE'].value_counts()

# Will manual add to PM_MAPPING later (will add later + will cross check later)
# EDMLO_Q_Check= DAY_EOD_File[DAY_EOD_File['Sub Business Unit'] == 'EDMLO_Q']

#For Checking purposes, Capturing False Case
False_Check= DAY_EOD_File[(DAY_EOD_File['TRUE/FALSE'] == 'FALSE')|(DAY_EOD_File['TRUE/FALSE'] == 'Match for trader only!')]
False_Check_2= HIERARCHY_DAY_PM[(HIERARCHY_DAY_PM['TRUE/FALSE'] == 'FALSE')|(HIERARCHY_DAY_PM['TRUE/FALSE'] == 'Match for trader only!')]
False_Check_Final = False_Check[["Transaction Type","Counterparty","Custodian Name","Strategy","BB Yellow Key","Trade Id","Instrument Type","PAGStagedVsDropcopy","Business Unit","Sub Business Unit","Order Creation User","HIERARCHY CHECKS","TRUE/FALSE"]]

# print success message if no false result for users
if False_Check_Final.shape[0]==0:
    print ('Hierarchy Check Pass!')
else:
    print ('Please Check for Hierarchy')

# Part 2: ACCT/FUND CHECKS 

# Map the FUND NAME with sub business unit
def ACCT_FUND_CHECK(row):
    SBU = row['Sub Business Unit']
    if SBU == 'EDMLO_Q': # Will manual add to PM_MAPPING later, will cross check later + delete this part after cross check
        return row['FundShortName']
    else:
        return lookup_dict_AF.get(row['Sub Business Unit'], "Unknown")

DAY_EOD_File['ACCT/FUND CHECKS'] = DAY_EOD_File.apply(ACCT_FUND_CHECK, axis=1)


#check whether look up value (FUND NAME) match with the FundShortName in original file
DAY_EOD_File['T/F'] = np.where(DAY_EOD_File['ACCT/FUND CHECKS'] == DAY_EOD_File['FundShortName'],'TRUE','FALSE' )

# Select the final result column into a separate dataframe
ACCT_DAY_PM = DAY_EOD_File[[ 'ACCT/FUND CHECKS' , 'T/F']]

#For Checking purposes, Check for any false value
column_counts_ACCT = ACCT_DAY_PM['T/F'].value_counts()

#For Checking purposes, Capturing False Case
False_Check_ACCT= DAY_EOD_File[DAY_EOD_File['T/F'] == 'FALSE']
False_Check_ACCT_Final = False_Check_ACCT[["Transaction Type","Counterparty","Custodian Name","Strategy","BB Yellow Key","Trade Id","Instrument Type","FundShortName","ACCT/FUND CHECKS","T/F"]]

if False_Check_ACCT_Final.shape[0]==0:
    print ('ACCT/FUND Check Pass!')
else:
    print ('Please Check for ACCT/FUND')









# Part 3: Borrow Agreement Checks

# Check if the transaction type is BuyToCover/ Sell Short while the instrument type is CASH (Non Swap + Non Future)
def BORROWING_REQUIRED(SWAP, TRAN_TYPE, INSTRUMENT,Custodian):
    if 'ISDA' in Custodian: # Exclude on those False but required case (contains ISDA)
        return "Not Required"
    else: 
        if TRAN_TYPE == "BuyToCover" and INSTRUMENT == "Equity":
            return "Required"
        elif TRAN_TYPE == "SellShort" and INSTRUMENT == "Equity":
            return "Required"
        else:
            return "Not Required"
        
# Check whether Borrow Agreement Description is empty 
DAY_EOD_File['Borrow Agreement Check'] = np.where(DAY_EOD_File['Borrow Agreement Description'].isna(),'FALSE','TRUE')

# Apply the function to check if the transaction type is BuyToCover/ Sell Short while the instrument type is CASH (Non Swap + Non Future)
DAY_EOD_File['Borrowing Agreement Requirement'] = DAY_EOD_File.apply(
    lambda row: BORROWING_REQUIRED(row['SWAP ORDER (NEW)'], row['Transaction Type'], row['Instrument Type'], row['Custodian Name']),
    axis=1
)



#For Checking purposes, Check for any false value
column_counts = DAY_EOD_File['Borrowing Agreement Requirement'].value_counts()
column_counts_BORROW = DAY_EOD_File['Borrow Agreement Check'].value_counts()

COUNT_BORROW_FALSE = column_counts_BORROW['FALSE']
COUNT_BORROW_TRUE =  column_counts_BORROW['TRUE']
COUNT_BORROW_REQUIED = column_counts['Required']
COUNT_BORROW_N_REQUIRED = column_counts['Not Required']


# Select the final result column into a separate dataframe
BORROW_DAY = DAY_EOD_File[[ 'Borrow Agreement Check' , 'Borrowing Agreement Requirement']]

#For Checking purposes, Capturing False Case
False_Check_BORROW= DAY_EOD_File[((DAY_EOD_File['Borrow Agreement Check'] == 'FALSE') & (DAY_EOD_File['Borrowing Agreement Requirement'] == 'Required'))
                                 |(DAY_EOD_File['Borrow Agreement Check'] == 'TRUE') & (DAY_EOD_File['Borrowing Agreement Requirement'] == 'Not Required')]

False_Check_BORROW_Final = False_Check_BORROW[["Transaction Type","Counterparty","Custodian Name","SWAP ORDER (NEW)","Strategy","BB Yellow Key","Trade Id","Instrument Type","Borrow Agreement Description","Borrow Agreement Check","Borrowing Agreement Requirement"]]

# print success message if no mismatch result for users 
if False_Check_BORROW_Final.shape[0]==0 and COUNT_BORROW_N_REQUIRED == COUNT_BORROW_FALSE and COUNT_BORROW_REQUIED == COUNT_BORROW_TRUE:
    print ('Borrow Agreement Check Pass!')
else:
    print ('Please Check for Borrow Agreement')


# Part 4: Financing Desciption Checks

#  check if the instrument type is Swap (contain ISDA)
def FINANCIAL_REQUIRED(SWAP, Custodian, INSTRUMENT):
    if 'ISDA' in Custodian and (INSTRUMENT == "Equity" or INSTRUMENT == "Listed Option" or INSTRUMENT == "Future"): 
        return "Required"
    else:
        return "Not Required"

# Check whether Financing Description is empty 
DAY_EOD_File['Financing Description Check'] = np.where(DAY_EOD_File['Financing Description'].isna(),'FALSE','TRUE')

# Apply the function to check if the instrument type is Swap (contain ISDA)
DAY_EOD_File['Financing Description Requirement'] = DAY_EOD_File.apply(
    lambda row: FINANCIAL_REQUIRED(row['SWAP ORDER (NEW)'], row['Custodian Name'], row['Instrument Type']),
    axis=1
)

#For Checking purposes, Check for any false value
column_counts_FINA_1 = DAY_EOD_File['Financing Description Check'].value_counts()
column_counts_FINA_2 = DAY_EOD_File['Financing Description Requirement'].value_counts()

COUNT_FINA_FALSE = column_counts_FINA_1['FALSE']
COUNT_FINA_TRUE =  column_counts_FINA_1['TRUE']
COUNT_FINA_REQUIED = column_counts_FINA_2['Required']
COUNT_FINA_N_REQUIRED = column_counts_FINA_2['Not Required']


FINANCIAL_DAY = DAY_EOD_File[[ 'Financing Description Check' , 'Financing Description Requirement']]

#For Checking purposes, Capturing False Case
False_Check_FINA =  DAY_EOD_File[((DAY_EOD_File['Financing Description Check'] == 'FALSE') & (DAY_EOD_File['Financing Description Requirement'] == 'Required'))
                                 |(DAY_EOD_File['Financing Description Check'] == 'TRUE') & (DAY_EOD_File['Financing Description Requirement'] == 'Not Required')]

False_Check_FINA_FINAL = False_Check_FINA[["Transaction Type","Counterparty","Custodian Name","SWAP ORDER (NEW)","Strategy","BB Yellow Key","Trade Id","Instrument Type","Financing Description","Financing Description Check","Financing Description Requirement"]]

# print success message if no mismatch result for users 
if False_Check_FINA_FINAL.shape[0]==0 and COUNT_FINA_N_REQUIRED == COUNT_FINA_FALSE and COUNT_FINA_REQUIED == COUNT_FINA_TRUE:
    print ('Financial Agreement Check Pass!')
else:
    print ('Please Check for Financial Agreement')




# Part 5: Check Preallocation
# Check if there's any preallocation (Not yet booked)
def check_preallocation(Custodian):
    if "preallocation" in Custodian or "Preallocation" in Custodian:
        return "not ok"
    else:
        return "ok"
    
DAY_EOD_File['Preallocation Check'] = DAY_EOD_File.apply(
      lambda row: check_preallocation(row['Custodian Name']),
      axis=1
  )  

PREALLOCATION_DAY = DAY_EOD_File[['Preallocation Check']]

#For Checking purposes, Check for any false value
column_counts_PRE = DAY_EOD_File['Preallocation Check'].value_counts()

False_Check_PRE= DAY_EOD_File[DAY_EOD_File['Preallocation Check'] == 'not ok']

False_Check_PRE_FINAL = False_Check_PRE[["Transaction Type","Counterparty","Custodian Name","Strategy","BB Yellow Key","Trade Id","Instrument Type","Preallocation Check"]]

# print success message if no false result for users 
if False_Check_PRE_FINAL.shape[0]==0:
    print ('Preallocation Check Pass!') 
else:
    print ('Please Check for Preallocation')

# Part 6: Mainbook Check

MAINBOOK_LIST = MAINBOOK_MAP['MAINBOOK FLAG'].tolist()

def MAIN_CHECK(Strategy): # Check whether the trade put into main strategy for specific pm
    if Strategy in MAINBOOK_LIST: #can edit in excel
        return "FLAG"
    else:
        return "GOOD"
    
DAY_EOD_File['Mainbook Check'] = DAY_EOD_File.apply(
      lambda row: MAIN_CHECK(row['Strategy']),
      axis=1
  )  

MAIN_DAY = DAY_EOD_File[['Mainbook Check']]

#For Checking purposes, Check for any false value
column_counts_MAIN = DAY_EOD_File['Mainbook Check'].value_counts()

False_Check_MAIN= DAY_EOD_File[DAY_EOD_File['Mainbook Check'] == 'FLAG']

False_Check_MAIN_Final = False_Check_MAIN[["Transaction Type","Counterparty","Custodian Name","Strategy","BB Yellow Key","Trade Id","Instrument Type","Mainbook Check"]]

# print success message if no false result for users 
if False_Check_MAIN_Final.shape[0]==0:
    print ('MAIN Book Check Pass!') 
else:
    print ('Please Check for MAIN Book')


# Part 7: Swap Check

def SWAP_CHECK(SWAP,Custodian): 
    if SWAP == True and "ISDA" in Custodian:
        return "GOOD"
    elif SWAP == False and "ISDA" not in Custodian:
         return "GOOD"
    elif math.isnan(SWAP):
        return "GOOD"
    else:
        return "FLAG"
    
DAY_EOD_File['Swap Check'] = DAY_EOD_File.apply(
      lambda row: SWAP_CHECK(row['SWAP ORDER (NEW)'],row['Custodian Name']),
      axis=1
  )  

SWAP_DAY = DAY_EOD_File[['Swap Check']]

#For Checking purposes, Check for any false value
column_counts_SWAP = DAY_EOD_File['Swap Check'].value_counts()

False_Check_SWAP= DAY_EOD_File[DAY_EOD_File['Swap Check'] == 'FLAG']

False_Check_SWAP_Final = False_Check_SWAP[["Transaction Type","Counterparty","SWAP ORDER (NEW)","Custodian Name","Strategy","BB Yellow Key","Trade Id","Instrument Type","Swap Check"]]

# print success message if no false result for users 
if False_Check_SWAP_Final.shape[0]==0:
    print ('Swap Check Pass!') 
else:
    print ('Please Check for Swap Check')





# Part 8: Final Part 
# Concatenate the dataframes horizontally (column-wise)
EOD_CHECK_DAY_FINAL = pd.concat([HIERARCHY_DAY_PM, ACCT_DAY_PM, BORROW_DAY, FINANCIAL_DAY, PREALLOCATION_DAY, MAIN_DAY, SWAP_DAY], axis=1)


EOD_DATA = {'Checks':['Hierarchy','Account'],'Errors':[False_Check_Final.shape[0],False_Check_ACCT_Final.shape[0]]}    
EOD_EMAIL = pd.DataFrame(EOD_DATA)


quant_list = []

for k in range(len(MAINBOOK_LIST)):
    quant_list.append('nil')


EOD_DATA_MAIN = {'MAINBOOK Check':MAINBOOK_LIST,
                 'Quantity':quant_list}

EOD_MAIN_EMAIL = pd.DataFrame(EOD_DATA_MAIN)

for index, row in EOD_MAIN_EMAIL.iterrows():
    Check_value = row['MAINBOOK Check']
    Error_Count = False_Check_MAIN_Final['Strategy'].eq(Check_value).sum()
    EOD_MAIN_EMAIL.at[index, 'Quantity'] = Error_Count
    

print(EOD_EMAIL)
print(EOD_MAIN_EMAIL)

#finding the no. of row for each false table (for formatting)
LENGTH_OVERNIGHT_F= 0
LENGTH_HIERARCHY_F= False_Check_Final.shape[0]
LENGTH_ACCT_F= False_Check_ACCT_Final.shape[0]
LENGTH_BORROW_F= False_Check_BORROW_Final.shape[0]
LENGTH_FINA_F= False_Check_FINA_FINAL.shape[0]
LENGTH_PRE_F= False_Check_PRE_FINAL.shape[0]
LENGTH_MAIN_F= False_Check_MAIN_Final.shape[0]
LENGTH_SWAP_F= False_Check_SWAP_Final.shape[0]
LENGTH_EMAIL_MAIN_F= EOD_MAIN_EMAIL.shape[0]

# yesterday_str for testing, will change to today_str
# Export_path= r"P:\Operations\Polymer - Middle Office\EOD Check\Christy\EOD_CHECK"+yesterday_str+"_DAY.csv"
# EOD_CHECK_DAY_FINAL.to_csv(Export_path, sheet_name)


# create a excel writer object
list_email = []

with pd.ExcelWriter("P:\Operations\Polymer - Middle Office\EOD Check\Christy\EOD_CHECK_"+yesterday_str_name+"_OVERNIGHT.xlsx") as writer:
   
    # use to_excel function and specify the sheet_name and index
    # to store the dataframe in specified sheet
    #DAY_EOD_File.to_excel(writer, sheet_name="TEMP", index=False)
    EOD_CHECK_DAY_FINAL.to_excel(writer, sheet_name="OVERALL", index=False)
    False_Check_OVERNIGHT_Final.to_excel(writer, sheet_name="OVERNIGHT", index=False)
    False_Check_Final.to_excel(writer, sheet_name="Hierarchy", index=False)
    False_Check_ACCT_Final.to_excel(writer, sheet_name="ACCT_FUND", index=False)
    False_Check_BORROW_Final.to_excel(writer, sheet_name="Borrow Agreement", index=False)
    False_Check_FINA_FINAL.to_excel(writer, sheet_name="Financial Agreement", index=False)
    False_Check_PRE_FINAL.to_excel(writer, sheet_name="Preallocation", index=False)
    False_Check_MAIN_Final.to_excel(writer, sheet_name="MAIN Book", index=False)
    False_Check_SWAP_Final.to_excel(writer, sheet_name="Swap", index=False)
    
    i = 5
    
    EOD_EMAIL.to_excel(writer, sheet_name="EMAIL SUMMARY", index=False, startrow=0)
    EOD_MAIN_EMAIL.to_excel(writer, sheet_name="EMAIL SUMMARY", index=False, startrow=5)
    
    
    if LENGTH_OVERNIGHT_F != 0:
        i = i+2
        ON=[["Overnight Check:"]]
        ON_df = pd.DataFrame(ON, columns =["Column1"], index=[1])
        list_email.append(ON_df)
        list_email.append(False_Check_OVERNIGHT_Final)
        False_Check_OVERNIGHT_Final.to_excel(writer, sheet_name="EMAIL SUMMARY", index=False, startrow=i+LENGTH_EMAIL_MAIN_F)
    if LENGTH_HIERARCHY_F != 0:
        i = i+2
        HC=[["Hierarchy Check:"]]
        HC_df = pd.DataFrame(HC, columns =["Column1"], index=[1])
        list_email.append(HC_df)
        list_email.append(False_Check_Final)
        False_Check_Final.to_excel(writer, sheet_name="EMAIL SUMMARY", index=False, startrow=i+LENGTH_EMAIL_MAIN_F+LENGTH_OVERNIGHT_F)
    if LENGTH_ACCT_F != 0:
        i = i+2
        AF=[["Acct/Fund Check:"]]
        AF_df = pd.DataFrame(AF, columns =["Column1"], index=[1])
        list_email.append(AF_df)
        list_email.append(False_Check_ACCT_Final)
        False_Check_ACCT_Final.to_excel(writer, sheet_name="EMAIL SUMMARY", index=False, startrow=i+LENGTH_EMAIL_MAIN_F+LENGTH_OVERNIGHT_F+LENGTH_HIERARCHY_F)
    if LENGTH_BORROW_F != 0:
        i = i+2
        BA=[["Borrow Agreement Check:"]]
        BA_df = pd.DataFrame(BA, columns =["Column1"], index=[1])
        list_email.append(BA_df)
        list_email.append(False_Check_BORROW_Final)
        False_Check_BORROW_Final.to_excel(writer, sheet_name="EMAIL SUMMARY", index=False, startrow=i+LENGTH_EMAIL_MAIN_F+LENGTH_OVERNIGHT_F+LENGTH_HIERARCHY_F+LENGTH_ACCT_F)
    if LENGTH_FINA_F != 0:
        i = i+2
        FA=[["Financial Agreement Check:"]]
        FA_df = pd.DataFrame(FA, columns =["Column1"], index=[1])
        list_email.append(FA_df)
        list_email.append(False_Check_FINA_FINAL)
        False_Check_FINA_FINAL.to_excel(writer, sheet_name="EMAIL SUMMARY", index=False, startrow=i+LENGTH_EMAIL_MAIN_F+LENGTH_OVERNIGHT_F+LENGTH_HIERARCHY_F+LENGTH_ACCT_F+LENGTH_BORROW_F)
    if LENGTH_PRE_F != 0:
        i = i+2
        Preallo=[["Preallocation Check:"]]
        Preallo_df = pd.DataFrame(Preallo, columns =["Column1"], index=[1])
        list_email.append(Preallo_df)
        list_email.append(False_Check_PRE_FINAL)
        False_Check_PRE_FINAL.to_excel(writer, sheet_name="EMAIL SUMMARY", index=False, startrow=i+LENGTH_EMAIL_MAIN_F+LENGTH_OVERNIGHT_F+LENGTH_HIERARCHY_F+LENGTH_ACCT_F+LENGTH_BORROW_F+LENGTH_FINA_F)
    if LENGTH_MAIN_F != 0:
        i = i+2
        MAIN_C=[["Mainbook Check:"]]
        MAIN_C_df = pd.DataFrame(MAIN_C, columns =["Column1"], index=[1])
        list_email.append(MAIN_C_df)
        list_email.append(False_Check_MAIN_Final)
        False_Check_MAIN_Final.to_excel(writer, sheet_name="EMAIL SUMMARY", index=False, startrow=i+LENGTH_EMAIL_MAIN_F+LENGTH_OVERNIGHT_F+LENGTH_HIERARCHY_F+LENGTH_ACCT_F+LENGTH_BORROW_F+LENGTH_FINA_F+LENGTH_PRE_F)        
    if LENGTH_SWAP_F != 0:
        i = i+2
        SWAP_C=[["Swap Check:"]]
        SWAP_C_df = pd.DataFrame(SWAP_C, columns =["Column1"], index=[1])
        list_email.append(SWAP_C_df)
        list_email.append(False_Check_SWAP_Final)
        False_Check_SWAP_Final.to_excel(writer, sheet_name="EMAIL SUMMARY", index=False, startrow=i+LENGTH_EMAIL_MAIN_F+LENGTH_OVERNIGHT_F+LENGTH_HIERARCHY_F+LENGTH_ACCT_F+LENGTH_BORROW_F+LENGTH_FINA_F+LENGTH_PRE_F+LENGTH_MAIN_F)        
 
# JUST try some fancy stuff        
import openpyxl
export_path ="P:\Operations\Polymer - Middle Office\EOD Check\Christy\EOD_CHECK_"+yesterday_str_name+"_OVERNIGHT.xlsx"
#export_path ="P:\Operations\Polymer - Middle Office\EOD Check\Christy\EOD_CHECK2023-08-07_DAY_CROSS.xlsx"

def colur_change(path):
    workbook = openpyxl.load_workbook(path)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        if sheet_name in ("Hierarchy","ACCT_FUND","Borrow Agreement","Financial Agreement","Preallocation","MAIN Book","Swap") and sheet.max_row>1:
            sheet.sheet_properties.tabColor = "DA9694"
            
    workbook.save(path)
    
colur_change(export_path)
        
        
#sending the email
import win32com.client as win32

dataframes = [EOD_EMAIL,EOD_MAIN_EMAIL,False_Check_Final,False_Check_ACCT_Final,False_Check_BORROW_Final,False_Check_FINA_FINAL,False_Check_PRE_FINAL,False_Check_MAIN_Final,False_Check_SWAP_Final]
always_eported_indices=[0,1]
always_exported_html=[dataframes[i].to_html(index=False) for i in always_eported_indices]

conditionally_exported_html=[]

for i in range(len(list_email)):
    if i%2!= 0:
        conditionally_exported_html.append(list_email[i].to_html(index=False))
    else:
        conditionally_exported_html.append(list_email[i].to_html(index=False, header=False))
    
combined_html = '<br><br>'.join(always_exported_html+conditionally_exported_html)
    
#html = EOD_EMAIL.to_html(index=False)+"<br><br>"+EOD_MAIN_EMAIL.to_html(index=False)

# if 'False_Check_Final' in list_email = []:
#     html = EOD_EMAIL.to_html(index=False)+"<br><br>"+EOD_MAIN_EMAIL.to_html(index=False)+"<br><br>"+EOD_MAIN_EMAIL.to_html(index=False)

# for dataframe in list_email:
#     html += dataframe.to_html(index=False)+"<br><br>"


text_file = open("P:\Operations\Polymer - Middle Office\EOD Check\Christy\TEST_1.htm",'w')
text_file.write(combined_html)
text_file.close()

outlook = win32.Dispatch('outlook.application')
mail = outlook.CreateItem(0)

mail.To = 'PolymerOps@polymercapital.com'
#mail.To = 'cmang@polymercapital.com;johnathanl@polymercapital.com>'
#mail.CC = PM_enail_mapping_CC_list

sub = "EMEA & Overnight EOD Check_" + yesterday_str_name
mail.Subject = sub
# mail.Body = 'Message body testing'

HtmlFile = open(r'P:\Operations\Polymer - Middle Office\EOD Check\Christy\TEST_1.htm', 'r')
source_code = HtmlFile.read()
HtmlFile.close()
mail.HTMLBody= source_code
mail.Display(True)
